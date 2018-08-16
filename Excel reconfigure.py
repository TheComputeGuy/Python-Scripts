import openpyxl as oxl
from datetime import datetime

wb = oxl.load_workbook("name.xlsx")

sheet = wb.active

count = 0

org_max = sheet.max_row

curr_max = org_max

for i in range(2, org_max+1):                       ##Iterating for the whole sheet. Starts from row 2 considering top row is headers
    val = sheet.cell(row=i, column=6).value         ##Column 6 consisted of the requisite data here
    if (", " in val):                               ##", " is the separator for Google Forms checkboxes response
        teams = val.split(', ')
        for j in range(len(teams)):
            for k in range(1, sheet.max_column+1):  ##Duplicate the current row to accomodate separate entries
                sheet.cell(row=curr_max, column=k).value = sheet.cell(
                    row=i, column=k).value
            sheet.cell(row=curr_max, column=6).value = teams[j]
            curr_max += 1
        count += 1
        for k in range(1, sheet.max_column+1):
            sheet.cell(row=i, column=k).value = None        ##Clear current row entries
print("Changed %d rows" % count)
wb.save("updatedName.xlsx")
